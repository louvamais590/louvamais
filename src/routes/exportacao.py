from flask import Blueprint, request, jsonify, make_response
from src.models.user import db
from src.models.escala import Escala
from src.models.escala_pessoa import EscalaPessoa
from src.models.pessoa import Pessoa
from datetime import datetime
import io
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.pdfgen import canvas

exportacao_bp = Blueprint('exportacao', __name__)

@exportacao_bp.route('/escalas/exportar-pdf', methods=['GET'])
def exportar_escalas_pdf():
    """Exporta as escalas em formato PDF"""
    try:
        # Parâmetros de filtro
        mes = request.args.get('mes', type=int)
        ano = request.args.get('ano', type=int)
        
        # Query base
        query = Escala.query.order_by(Escala.data)
        
        # Aplicar filtros se fornecidos
        if mes and ano:
            query = query.filter(
                db.extract('month', Escala.data) == mes,
                db.extract('year', Escala.data) == ano
            )
        elif ano:
            query = query.filter(db.extract('year', Escala.data) == ano)
        
        escalas = query.all()
        
        if not escalas:
            return jsonify({
                'success': False,
                'error': 'Nenhuma escala encontrada para exportar'
            }), 404
        
        # Criar PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
        
        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1,  # Center
            textColor=colors.HexColor('#667eea')
        )
        
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=12,
            textColor=colors.HexColor('#4a5568')
        )
        
        # Conteúdo do PDF
        story = []
        
        # Título
        title = "Escala do Grupo de Oração"
        if mes and ano:
            meses = ['', 'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
                    'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
            title += f" - {meses[mes]} {ano}"
        elif ano:
            title += f" - {ano}"
        
        story.append(Paragraph(title, title_style))
        story.append(Spacer(1, 20))
        
        # Agrupar escalas por mês
        escalas_por_mes = {}
        for escala in escalas:
            mes_ano = escala.data.strftime('%Y-%m')
            if mes_ano not in escalas_por_mes:
                escalas_por_mes[mes_ano] = []
            escalas_por_mes[mes_ano].append(escala)
        
        # Gerar conteúdo para cada mês
        for mes_ano, escalas_mes in escalas_por_mes.items():
            # Título do mês
            data_mes = datetime.strptime(mes_ano, '%Y-%m')
            titulo_mes = data_mes.strftime('%B %Y').title()
            story.append(Paragraph(titulo_mes, subtitle_style))
            
            # Criar tabela para o mês
            data_tabela = []
            
            # Cabeçalho da tabela
            if any(escala.dia_semana.lower().find('terça') != -1 for escala in escalas_mes):
                # Tem terças-feiras
                headers = ['Data', 'Dia', 'Pregação', 'Músicos', 'Condução/Oração', 'Acolhida', 'Abastecimento']
            else:
                # Só quartas-feiras
                headers = ['Data', 'Dia', 'Abastecimento']
            
            data_tabela.append(headers)
            
            # Dados das escalas
            for escala in escalas_mes:
                linha = [
                    escala.data.strftime('%d/%m'),
                    escala.dia_semana
                ]
                
                if escala.dia_semana.lower().find('terça') != -1:
                    # Terça-feira
                    linha.extend([
                        escala.pregacao_display or '-',
                        escala.musicos_display or '-',
                        escala.conducao_animacao_display or '-',
                        escala.acolhida_display or '-',
                        '-'  # Abastecimento vazio para terças
                    ])
                else:
                    # Quarta-feira
                    if len(headers) > 3:
                        # Tabela completa
                        linha.extend(['-', '-', '-', '-', escala.abastecimento_display or '-'])
                    else:
                        # Tabela só de quartas
                        linha.append(escala.abastecimento_display or '-')
                
                data_tabela.append(linha)
            
            # Criar e estilizar tabela
            tabela = Table(data_tabela)
            tabela.setStyle(TableStyle([
                # Cabeçalho
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                
                # Dados
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
                
                # Bordas
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0')),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ]))
            
            story.append(tabela)
            story.append(Spacer(1, 20))
        
        # Rodapé
        rodape_style = ParagraphStyle(
            'Rodape',
            parent=styles['Normal'],
            fontSize=8,
            alignment=1,
            textColor=colors.HexColor('#64748b')
        )
        
        data_geracao = datetime.now().strftime('%d/%m/%Y às %H:%M')
        story.append(Spacer(1, 30))
        story.append(Paragraph(f"Relatório gerado em {data_geracao}", rodape_style))
        
        # Gerar PDF
        doc.build(story)
        
        # Preparar resposta
        buffer.seek(0)
        pdf_data = buffer.getvalue()
        buffer.close()
        
        # Nome do arquivo
        nome_arquivo = "escala_grupo_oracao"
        if mes and ano:
            nome_arquivo += f"_{ano}_{mes:02d}"
        elif ano:
            nome_arquivo += f"_{ano}"
        nome_arquivo += ".pdf"
        
        response = make_response(pdf_data)
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename="{nome_arquivo}"'
        
        return response
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Erro ao gerar PDF: {str(e)}'
        }), 500

@exportacao_bp.route('/escalas/exportar-excel', methods=['GET'])
def exportar_escalas_excel():
    """Exporta as escalas em formato Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # Parâmetros de filtro
        mes = request.args.get('mes', type=int)
        ano = request.args.get('ano', type=int)
        
        # Query base
        query = Escala.query.order_by(Escala.data)
        
        # Aplicar filtros se fornecidos
        if mes and ano:
            query = query.filter(
                db.extract('month', Escala.data) == mes,
                db.extract('year', Escala.data) == ano
            )
        elif ano:
            query = query.filter(db.extract('year', Escala.data) == ano)
        
        escalas = query.all()
        
        if not escalas:
            return jsonify({
                'success': False,
                'error': 'Nenhuma escala encontrada para exportar'
            }), 404
        
        # Criar workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Escala Grupo de Oração"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Cabeçalhos
        headers = ['Data', 'Dia da Semana', 'Pregação', 'Equipe Músicos', 
                  'Condução de Animação/Oração', 'Acolhida', 'Responsável Condução Abastecimento']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        
        # Dados
        for row, escala in enumerate(escalas, 2):
            ws.cell(row=row, column=1, value=escala.data.strftime('%d/%m/%Y')).border = border
            ws.cell(row=row, column=2, value=escala.dia_semana).border = border
            
            if escala.dia_semana.lower().find('terça') != -1:
                ws.cell(row=row, column=3, value=escala.pregacao_display or '').border = border
                ws.cell(row=row, column=4, value=escala.musicos_display or '').border = border
                ws.cell(row=row, column=5, value=escala.conducao_animacao_display or '').border = border
                ws.cell(row=row, column=6, value=escala.acolhida_display or '').border = border
                ws.cell(row=row, column=7, value='N/A').border = border
            else:
                ws.cell(row=row, column=3, value='N/A').border = border
                ws.cell(row=row, column=4, value='N/A').border = border
                ws.cell(row=row, column=5, value='N/A').border = border
                ws.cell(row=row, column=6, value='N/A').border = border
                ws.cell(row=row, column=7, value=escala.abastecimento_display or '').border = border
        
        # Ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Salvar em buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Nome do arquivo
        nome_arquivo = "escala_grupo_oracao"
        if mes and ano:
            nome_arquivo += f"_{ano}_{mes:02d}"
        elif ano:
            nome_arquivo += f"_{ano}"
        nome_arquivo += ".xlsx"
        
        response = make_response(buffer.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename="{nome_arquivo}"'
        
        return response
        
    except ImportError:
        return jsonify({
            'success': False,
            'error': 'Biblioteca openpyxl não está disponível'
        }), 500
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Erro ao gerar Excel: {str(e)}'
        }), 500

